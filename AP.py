import streamlit as st
import imaplib
import email
from bs4 import BeautifulSoup
import re
import pandas as pd
import base64
from PIL import Image
import io
import docx2txt
import openpyxl
import PyPDF2

# Streamlit app title
st.title("Automate2Excel: Simplified Data Transfer")

# Create input fields for the user and password
user = st.text_input("Enter your email address")
password = st.text_input("Enter your email password", type="password")

# Create input field for the email address to search for
search_email = st.text_input("Enter the email address to search for")

# Create dropdown for document type
document_type = st.selectbox("Select the type of document", ["HTML", "Image", "Word", "Excel", "PDF"])

# Create date picker for selecting the date
selected_date = st.date_input("Select the date of the mail")

# Function to extract information from HTML content
def extract_info_from_html(html_content):
    soup = BeautifulSoup(html_content, 'html.parser')
    
    info = {
        "Name": None,
        "Email": None,
        "Workshop Detail": None,
        "Date": None,
        "Mobile No.": None
    }

    name_element = soup.find(string=re.compile(r'Name', re.IGNORECASE))
    if name_element:
        next_td = name_element.find_next('td')
        info["Name"] = next_td.get_text().strip() if next_td else None

    email_element = soup.find(string=re.compile(r'Email', re.IGNORECASE))
    if email_element:
        next_td = email_element.find_next('td')
        info["Email"] = next_td.get_text().strip() if next_td else None

    workshop_element = soup.find(string=re.compile(r'Workshop Detail', re.IGNORECASE))
    if workshop_element:
        next_td = workshop_element.find_next('td')
        info["Workshop Detail"] = next_td.get_text().strip() if next_td else None

    date_element = soup.find(string=re.compile(r'Date', re.IGNORECASE))
    if date_element:
        next_td = date_element.find_next('td')
        info["Date"] = next_td.get_text().strip() if next_td else None

    mobile_element = soup.find(string=re.compile(r'Mobile No\.', re.IGNORECASE))
    if mobile_element:
        next_td = mobile_element.find_next('td')
        info["Mobile No."] = next_td.get_text().strip() if next_td else None

    return info

# Function to extract text from an image
def extract_text_from_image(image_bytes):
    image = Image.open(io.BytesIO(image_bytes))
    text = pytesseract.image_to_string(image)
    return text

# Function to extract text from Word document
def extract_text_from_word(docx_bytes):
    text = docx2txt.process(io.BytesIO(docx_bytes))
    return text

# Function to extract text from Excel file
def extract_text_from_excel(excel_bytes):
    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes))
    text = []
    for sheet in wb.sheetnames:
        for row in wb[sheet].iter_rows(values_only=True):
            text.extend(row)
    return text

# Function to extract text from PDF
def extract_text_from_pdf(pdf_bytes):
    pdf_reader = PyPDF2.PdfReader(io.BytesIO(pdf_bytes))
    text = ""
    for page_num in range(len(pdf_reader.pages)):
        text += pdf_reader.pages[page_num].extract_text()
    return text

# Initialize mail_id_list outside the try block
mail_id_list = []

if st.button("Fetch and Generate Extracted Document"):
    try:
        # URL for IMAP connection
        imap_url = 'imap.gmail.com'

        # Connection with GMAIL using SSL
        my_mail = imaplib.IMAP4_SSL(imap_url)

        # Log in using user and password
        my_mail.login(user, password)

        # Select the Inbox to fetch messages
        my_mail.select('inbox')

        # Define the key and value for email search
        key = 'FROM'
        value = search_email  # Use the user-inputted email address to search
        _, data = my_mail.search(None, key, value)

        mail_id_list = data[0].split()

        info_list = []

        # Iterate through messages and extract information based on the selected document type
        for num in mail_id_list:
            typ, data = my_mail.fetch(num, '(RFC822)')
            msg = email.message_from_bytes(data[0][1])

            for part in msg.walk():
                if part.get_content_type() == 'text/html':
                    html_content = part.get_payload(decode=True).decode('utf-8')
                    info = extract_info_from_html(html_content)

                    # Extract and add the received date
                    date = msg["Date"]
                    info["Received Date"] = date

                    info_list.append(info)

                elif document_type == "Image" and part.get_content_type().startswith("image"):
                    # Call the function to extract text from image
                    image_text = extract_text_from_image(part.get_payload(decode=True))
                    st.text(f"Extracted Text from Image: {image_text}")

                elif document_type == "Word" and part.get_content_type().startswith("application/msword"):
                    # Call the function to extract text from Word document
                    word_text = extract_text_from_word(part.get_payload(decode=True))
                    st.text(f"Extracted Text from Word: {word_text}")

                elif document_type == "Excel" and part.get_content_type().startswith("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"):
                    # Call the function to extract text from Excel file
                    excel_text = extract_text_from_excel(part.get_payload(decode=True))
                    st.text(f"Extracted Text from Excel: {excel_text}")

                elif document_type == "PDF" and part.get_content_type().startswith("application/pdf"):
                    # Call the function to extract text from PDF
                    pdf_text = extract_text_from_pdf(part.get_payload(decode=True))
                    st.text(f"Extracted Text from PDF: {pdf_text}")

        # If the selected document type is HTML, create a DataFrame from the info_list
        if document_type == "HTML":
            df = pd.DataFrame(info_list)
            st.write("Data extracted from emails:")
            st.write(df)

        st.success("Extraction completed.")

    except Exception as e:
        st.error(f"Error: {e}")
else:
    st.warning("Click the 'Fetch and Generate Extracted Document' button to retrieve and process emails.")
